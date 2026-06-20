"""DGII certification Excel import and scenario planning."""

from __future__ import annotations

import hashlib
import logging
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.etree import ElementTree as ET

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import DataError, transaction
from django.utils import timezone

from facturacion.models import (
    Company,
    DGIICertificationEvent,
    DGIICertificationItem,
    DGIICertificationPlan,
)


logger = logging.getLogger(__name__)
MAX_AMOUNT_ABS = Decimal('999999999999.99')
SUPPORTED_ECF_TYPES = {'31', '32', '33', '34', '41', '43', '44', '45', '46', '47', 'RFCE'}
CERTIFICATION_XML_SUPPORTED_TYPES = {'31', '32', '33', '34', '41', '43', '44', '45', '46', '47'}
ENCF_RE = re.compile(r'\bE(31|32|33|34|41|43|44|45|46|47)\d{6,12}\b', re.IGNORECASE)
ENCF_FRAGMENT_RE = re.compile(r'E(31|32|33|34|41|43|44|45|46|47)\d{6,12}', re.IGNORECASE)
RNC_RE = re.compile(r'\b\d{3}-?\d{5}-?\d{1}\b|\b\d{3}-?\d{7}-?\d{1}\b')


TYPE_TEXT_HINTS = {
    '31': ['factura credito fiscal', 'credito fiscal', 'e31'],
    '32': ['factura consumo', 'consumo', 'e32'],
    '33': ['nota debito', 'debito', 'e33'],
    '34': ['nota credito', 'credito', 'e34'],
    '41': ['compras', 'compra', 'e41'],
    '43': ['gastos menores', 'gasto menor', 'e43'],
    '44': ['regimenes especiales', 'regimen especial', 'e44'],
    '45': ['gubernamental', 'gobierno', 'e45'],
    '46': ['exportaciones', 'exportacion', 'e46'],
    '47': ['pagos exterior', 'pago exterior', 'e47'],
    'RFCE': ['rfce', 'resumen factura consumo', 'resumen facturas consumo'],
}

HEADER_HINTS = {
    'encf': ['encf', 'e-ncf'],
    'amount': ['monto', 'total', 'valor', 'importe'],
    'receiver_rnc': ['rnc receptor', 'rncreceptor', 'rnc comprador', 'rnccomprador', 'rnc cliente', 'rnccliente'],
    'receiver_name': ['nombre receptor', 'nombrereceptor', 'razon social receptor', 'razonsocialreceptor', 'razon social comprador', 'razonsocialcomprador', 'cliente'],
    'observations': ['observacion', 'observaciones', 'comentario', 'nota'],
    'document_type': ['tipo comprobante', 'tipocomprobante', 'tipoecf', 'tipo e-cf', 'tipo ecf'],
    'dgii_group': ['grupo dgii', 'grupo'],
}


@dataclass(frozen=True)
class DetectedCertificationItem:
    ecf_type: str
    dgii_group: int
    encf: str
    document_type: str
    amount: Decimal | None
    receiver_rnc: str
    receiver_name: str
    observations: str
    source_sheet: str
    source_row: int
    raw_data: dict


class DGIICertificationExcelImporter:
    """Import the DGII certification workbook into a tenant-scoped plan."""

    max_size_bytes = 5 * 1024 * 1024

    def import_workbook(self, *, uploaded_file, company: Company, user=None) -> DGIICertificationPlan:
        filename = uploaded_file.name or 'set-dgii.xlsx'
        suffix = Path(filename).suffix.lower()
        content = uploaded_file.read()

        if not content:
            raise ValueError('Debe cargar el Excel entregado por DGII.')
        if len(content) > self.max_size_bytes:
            raise ValueError('El Excel DGII no puede exceder 5 MB.')
        if suffix not in {'.xlsx', '.xls'}:
            raise ValueError('El archivo debe ser .xlsx o .xls.')

        try:
            detected_items = self._detect_items(content, suffix)
        except Exception as exc:
            DGIICertificationEvent.objects.create(
                company=company,
                event_type=DGIICertificationEvent.EVENT_IMPORT_ERROR,
                message='No fue posible analizar el Excel DGII.',
                payload={'filename': filename, 'error': str(exc)},
                created_by=user,
            )
            raise ValueError(f'No fue posible analizar el Excel DGII: {exc}') from exc

        if not detected_items:
            DGIICertificationEvent.objects.create(
                company=company,
                event_type=DGIICertificationEvent.EVENT_IMPORT_ERROR,
                message='No se detectaron escenarios e-CF en el Excel DGII.',
                payload={'filename': filename},
                created_by=user,
            )
            raise ValueError('No se detectaron escenarios e-CF en el Excel DGII.')

        file_hash = hashlib.sha256(content).hexdigest()
        group_counts = self._group_counts(detected_items)

        try:
            with transaction.atomic():
                plan = DGIICertificationPlan.objects.create(
                    company=company,
                    source_filename=filename,
                    file_sha256=file_hash,
                    imported_by=user,
                    total_items=len(detected_items),
                    group_counts=group_counts,
                )
                DGIICertificationEvent.objects.create(
                    company=company,
                    plan=plan,
                    event_type=DGIICertificationEvent.EVENT_EXCEL_IMPORTED,
                    message='Excel DGII importado.',
                    payload={'filename': filename, 'sha256': file_hash, 'items': len(detected_items)},
                    created_by=user,
                )
                DGIICertificationEvent.objects.create(
                    company=company,
                    plan=plan,
                    event_type=DGIICertificationEvent.EVENT_PLAN_CREATED,
                    message='Plan de certificacion DGII creado.',
                    payload={'group_counts': group_counts},
                    created_by=user,
                )

                item_objects = []
                for detected in detected_items:
                    self._validate_detected_item(detected)
                    logger.info(
                        "DGII certification item detected before persistence",
                        extra={
                            "sheet_name": detected.source_sheet,
                            "row_number": detected.source_row,
                            "document_type": detected.document_type,
                            "encf": detected.encf,
                            "amount": str(detected.amount) if detected.amount is not None else None,
                            "rnc": detected.receiver_rnc,
                            "customer_name": detected.receiver_name,
                            "group_number": detected.dgii_group,
                        },
                    )
                    item = DGIICertificationItem.objects.create(
                        plan=plan,
                        company=company,
                        ecf_type=detected.ecf_type,
                        dgii_group=detected.dgii_group,
                        encf=detected.encf,
                        document_type=detected.document_type,
                        amount=detected.amount,
                        receiver_rnc=detected.receiver_rnc,
                        receiver_name=detected.receiver_name,
                        observations=detected.observations,
                        source_sheet=detected.source_sheet,
                        source_row=detected.source_row,
                        raw_data=detected.raw_data,
                    )
                    item_objects.append(item)
                    DGIICertificationEvent.objects.create(
                        company=company,
                        plan=plan,
                        item=item,
                        event_type=DGIICertificationEvent.EVENT_ITEM_DETECTED,
                        message=f'Item DGII detectado: {detected.ecf_type}.',
                        payload={
                            'ecf_type': detected.ecf_type,
                            'dgii_group': detected.dgii_group,
                            'sheet': detected.source_sheet,
                            'row': detected.source_row,
                        },
                        created_by=user,
                    )
        except (DataError, ValueError, InvalidOperation) as exc:
            DGIICertificationEvent.objects.create(
                company=company,
                event_type=DGIICertificationEvent.EVENT_IMPORT_ERROR,
                message='No se pudo importar el Excel DGII.',
                payload={'filename': filename, 'error': str(exc)},
                created_by=user,
            )
            raise ValueError(str(exc)) from exc

        return plan

    def _validate_detected_item(self, detected: DetectedCertificationItem) -> None:
        if detected.amount is None:
            return
        if abs(detected.amount) > MAX_AMOUNT_ABS:
            raise ValueError(
                "Monto fuera de rango para plan DGII "
                f"(hoja={detected.source_sheet}, fila={detected.source_row}, "
                f"valor={detected.amount}, campo=amount)."
            )

    def _detect_items(self, content: bytes, suffix: str) -> list[DetectedCertificationItem]:
        if suffix == '.xls':
            return self._detect_xls_items(content)
        return self._detect_xlsx_items(content)

    def _detect_xlsx_items(self, content: bytes) -> list[DetectedCertificationItem]:
        from io import BytesIO

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        detected: list[DetectedCertificationItem] = []
        seen_sources: set[tuple[str, int]] = set()

        for worksheet in workbook.worksheets:
            header_map: dict[int, str] = {}
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = list(row or [])
                normalized_values = [_normalize_text(value) for value in values]
                maybe_headers = self._header_map(normalized_values)
                if len(maybe_headers) >= 2 and not self._row_has_ecf_signal(values, normalized_values):
                    header_map = maybe_headers
                    continue

                item = self._detect_row(
                    values=values,
                    normalized_values=normalized_values,
                    header_map=header_map,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
                if item and (item.source_sheet, item.source_row) not in seen_sources:
                    detected.append(item)
                    seen_sources.add((item.source_sheet, item.source_row))

        return detected

    def _detect_xls_items(self, _content: bytes) -> list[DetectedCertificationItem]:
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise ValueError('Los archivos .xls requieren xlrd instalado; convierta el set DGII a .xlsx.') from exc

        raise ValueError('El motor .xls legacy no esta habilitado para este importador.')

    def _detect_row(
        self,
        *,
        values: list,
        normalized_values: list[str],
        header_map: dict[int, str],
        sheet_name: str,
        row_number: int,
    ) -> DetectedCertificationItem | None:
        row_text = ' '.join(value for value in normalized_values if value)
        if not row_text:
            return None

        field_values = self._field_values(values, normalized_values, header_map)
        encf = field_values.get('encf') or self._extract_encf(row_text)
        ecf_type = self._detect_ecf_type(row_text, encf, sheet_name, field_values.get('document_type', ''))
        if ecf_type not in SUPPORTED_ECF_TYPES:
            return None

        amount = self._extract_amount(values, normalized_values, header_map, field_values)
        receiver_rnc = field_values.get('receiver_rnc') or self._extract_rnc(row_text, encf)
        receiver_name = _clean_receiver_name(field_values.get('receiver_name')) or self._extract_name(values, normalized_values, header_map)
        observations = field_values.get('observations') or ''
        document_type = field_values.get('document_type') or self._document_type_label(ecf_type)
        dgii_group = self._classify_group(ecf_type, amount)
        raw_data = {
            f'col_{index + 1}': self._serialize_cell(value)
            for index, value in enumerate(values)
            if value not in (None, '')
        }

        return DetectedCertificationItem(
            ecf_type=ecf_type,
            dgii_group=dgii_group,
            encf=encf,
            document_type=document_type,
            amount=amount,
            receiver_rnc=receiver_rnc,
            receiver_name=receiver_name,
            observations=observations,
            source_sheet=sheet_name,
            source_row=row_number,
            raw_data=raw_data,
        )

    def _header_map(self, normalized_values: list[str]) -> dict[int, str]:
        headers: dict[int, str] = {}
        for index, value in enumerate(normalized_values):
            if not value:
                continue
            for field, hints in HEADER_HINTS.items():
                if any(hint in value for hint in hints):
                    headers[index] = field
                    break
        return headers

    def _row_has_ecf_signal(self, values: list, normalized_values: list[str]) -> bool:
        corpus = ' '.join(value for value in normalized_values if value)
        if ENCF_FRAGMENT_RE.search(corpus.upper()):
            return True
        return any(value.strip().upper() == 'RFCE' for value in normalized_values) or any(
            str(value).strip() in {'31', '32', '33', '34', '41', '43', '44', '45', '46', '47'}
            for value in values
            if value is not None
        )

    def _field_values(self, values: list, normalized_values: list[str], header_map: dict[int, str]) -> dict[str, str]:
        fields: dict[str, str] = {}
        for index, field in header_map.items():
            if index >= len(values):
                continue
            value = values[index]
            if value in (None, ''):
                continue
            text = str(value).strip()
            if field == 'receiver_rnc':
                text = _normalize_digits(text)
            elif field == 'encf':
                extracted = self._extract_encf(_normalize_text(text))
                if not extracted or field in fields:
                    continue
                text = extracted
            elif field == 'amount':
                continue
            elif field in fields:
                continue
            fields[field] = text
        return fields

    def _detect_ecf_type(self, row_text: str, encf: str, sheet_name: str, document_type: str) -> str | None:
        if self._is_rfce_context(row_text, sheet_name, document_type):
            return 'RFCE'

        if encf:
            match = ENCF_FRAGMENT_RE.search(encf)
            if match:
                return match.group(1)

        corpus = f'{row_text} {_normalize_text(sheet_name)} {_normalize_text(document_type)}'
        for ecf_type, hints in TYPE_TEXT_HINTS.items():
            if any(hint in corpus for hint in hints):
                return ecf_type
        return None

    def _extract_encf(self, text: str) -> str:
        match = ENCF_FRAGMENT_RE.search(text.upper())
        return match.group(0).upper() if match else ''

    def _extract_rnc(self, row_text: str, encf: str) -> str:
        for match in RNC_RE.finditer(row_text):
            digits = _normalize_digits(match.group(0))
            if len(digits) in {9, 11} and digits not in encf:
                return digits
        return ''

    def _extract_amount(self, values: list, normalized_values: list[str], header_map: dict[int, str], field_values: dict[str, str]) -> Decimal | None:
        amount_indexes = [index for index, field in header_map.items() if field == 'amount']
        header_candidates = []
        for index in amount_indexes:
            if index < len(values):
                parsed = _parse_decimal(values[index])
                if parsed is not None:
                    header_candidates.append(parsed)
        if header_candidates:
            return max(header_candidates)

        candidates = [
            parsed
            for value, normalized in zip(values, normalized_values)
            if not ENCF_FRAGMENT_RE.search(normalized.upper())
            for parsed in [_parse_decimal(value)]
            if parsed is not None
        ]
        return max(candidates) if candidates else None

    def _extract_name(self, values: list, normalized_values: list[str], header_map: dict[int, str]) -> str:
        used_header_indexes = set(header_map.keys())
        for index, value in enumerate(values):
            if index in used_header_indexes or value in (None, ''):
                continue
            text = str(value).strip()
            normalized = normalized_values[index]
            text = _clean_receiver_name(text)
            if not text:
                continue
            normalized = _normalize_text(text)
            if ENCF_FRAGMENT_RE.search(normalized.upper()) or RNC_RE.search(normalized):
                continue
            if _parse_decimal(value) is not None:
                continue
            if any(hint in normalized for hints in TYPE_TEXT_HINTS.values() for hint in hints):
                continue
            return text[:180]
        return ''

    def _is_rfce_context(self, row_text: str, sheet_name: str, document_type: str) -> bool:
        corpus = f'{_normalize_text(sheet_name)} {row_text} {_normalize_text(document_type)}'
        return any(hint in corpus for hint in TYPE_TEXT_HINTS['RFCE'])

    def _classify_group(self, ecf_type: str, amount: Decimal | None) -> int:
        if ecf_type in {'31', '41', '43', '44', '45', '46', '47'}:
            return 1
        if ecf_type == '32':
            if amount is not None and amount < Decimal('250000'):
                return 4
            return 1
        if ecf_type in {'33', '34'}:
            return 2
        if ecf_type == 'RFCE':
            return 3
        return 1

    def _group_counts(self, items: list[DetectedCertificationItem]) -> dict:
        counts = {'1': 0, '2': 0, '3': 0, '4': 0}
        for item in items:
            counts[str(item.dgii_group)] = counts.get(str(item.dgii_group), 0) + 1
        return counts

    def _document_type_label(self, ecf_type: str) -> str:
        return dict(DGIICertificationItem.ECF_TYPE_CHOICES).get(ecf_type, ecf_type)

    def _serialize_cell(self, value):
        if isinstance(value, Decimal):
            return str(value)
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        return value


def _normalize_text(value) -> str:
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'\s+', ' ', text)


def _normalize_digits(value) -> str:
    return re.sub(r'\D+', '', str(value or ''))


def _clean_receiver_name(value) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    normalized = _normalize_text(text)
    if normalized in {'#e', 'e', 'n/a', 'na', 'null', 'none', 'no aplica'}:
        return ''
    if len(normalized) < 3:
        return ''
    if not re.search(r'[A-Za-z]', text):
        return ''
    alpha_count = sum(1 for char in normalized if char.isalpha())
    if alpha_count < 3:
        return ''
    return text


def _parse_decimal(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            return None

    text = str(value).strip()
    if not re.search(r'\d', text):
        return None
    if re.search(r'[A-Za-z]', text):
        return None
    cleaned = re.sub(r'[^\d,.\-]', '', text)
    if cleaned.count(',') == 1 and cleaned.count('.') > 1:
        cleaned = cleaned.replace('.', '').replace(',', '.')
    elif cleaned.count(',') == 1 and cleaned.count('.') == 0:
        cleaned = cleaned.replace(',', '.')
    else:
        cleaned = cleaned.replace(',', '')
    try:
        return Decimal(cleaned).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return None


class DGIICertificationXMLGenerator:
    """Build isolated DGII certification scenarios without touching productive fiscal flows."""

    unsupported_rfce_message = 'Generador RFCE pendiente de implementación'

    def generate_item(self, *, item: DGIICertificationItem, user=None) -> DGIICertificationItem:
        if item.ecf_type == 'RFCE':
            return self._mark_generation_error(item=item, user=user, error=self.unsupported_rfce_message)
        if item.ecf_type not in CERTIFICATION_XML_SUPPORTED_TYPES:
            return self._mark_generation_error(
                item=item,
                user=user,
                error=f'Generador no disponible para tipo {item.ecf_type}.',
            )

        xml_content = self._build_xml(item)
        encoded = xml_content.encode('utf-8')
        digest = hashlib.sha256(encoded).hexdigest()
        filename = self._filename_for(item)
        storage_path = default_storage.save(filename, ContentFile(encoded))

        item.generated_xml_path = storage_path
        item.generated_xml_hash = digest
        item.generated_at = timezone.now()
        item.generation_error = ''
        item.status = DGIICertificationItem.STATUS_GENERATED
        item.save(update_fields=[
            'generated_xml_path',
            'generated_xml_hash',
            'generated_at',
            'generation_error',
            'status',
            'updated_at',
        ])
        DGIICertificationEvent.objects.create(
            company=item.company,
            plan=item.plan,
            item=item,
            event_type=DGIICertificationEvent.EVENT_XML_GENERATED,
            message='Escenario de certificacion DGII generado.',
            payload={
                'ecf_type': item.ecf_type,
                'dgii_group': item.dgii_group,
                'path': storage_path,
                'sha256': digest,
            },
            created_by=user,
        )
        return item

    def generate_group(self, *, plan: DGIICertificationPlan, group_number: int, user=None) -> dict:
        queryset = plan.items.filter(dgii_group=group_number, status=DGIICertificationItem.STATUS_PENDING).order_by(
            'source_sheet',
            'source_row',
        )
        generated = 0
        failed = 0
        errors = []
        for item in queryset:
            updated_item = self.generate_item(item=item, user=user)
            if updated_item.status == DGIICertificationItem.STATUS_GENERATED:
                generated += 1
            else:
                failed += 1
                errors.append({
                    'item_id': item.id,
                    'ecf_type': item.ecf_type,
                    'source_sheet': item.source_sheet,
                    'source_row': item.source_row,
                    'error': updated_item.generation_error,
                })
        return {'generated': generated, 'failed': failed, 'errors': errors}

    def _mark_generation_error(self, *, item: DGIICertificationItem, user=None, error: str) -> DGIICertificationItem:
        item.status = DGIICertificationItem.STATUS_GENERATION_ERROR
        item.generation_error = error
        item.save(update_fields=['status', 'generation_error', 'updated_at'])
        DGIICertificationEvent.objects.create(
            company=item.company,
            plan=item.plan,
            item=item,
            event_type=DGIICertificationEvent.EVENT_XML_GENERATION_ERROR,
            message=error,
            payload={'ecf_type': item.ecf_type, 'dgii_group': item.dgii_group},
            created_by=user,
        )
        return item

    def _build_xml(self, item: DGIICertificationItem) -> str:
        root = ET.Element('EscenarioCertificacionDGII')
        self._set_text(root, 'ecf_type', item.ecf_type)
        self._set_text(root, 'encf', item.encf)
        if item.amount is not None:
            self._set_text(root, 'amount', f'{item.amount:.2f}')
        self._set_text(root, 'receiver_rnc', item.receiver_rnc)
        self._set_text(root, 'receiver_name', item.receiver_name)
        self._set_text(root, 'observations', item.observations)
        self._set_text(root, 'group_number', str(item.dgii_group))
        self._set_text(root, 'source_sheet', item.source_sheet)
        self._set_text(root, 'source_row', str(item.source_row))

        ET.indent(root, space='  ')
        return ET.tostring(root, encoding='unicode', xml_declaration=True)

    def _set_text(self, parent, tag: str, value) -> None:
        cleaned = _clean_scenario_value(value)
        if not cleaned:
            return
        element = ET.SubElement(parent, tag)
        element.text = cleaned

    def _filename_for(self, item: DGIICertificationItem) -> str:
        encf = item.encf or f'fila-{item.source_row}'
        return (
            f'dgii_certification/company-{item.company_id}/plan-{item.plan_id}/'
            f'grupo-{item.dgii_group}/{item.ecf_type}-{encf}-item-{item.id}.xml'
        )

def _clean_scenario_value(value) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    normalized = _normalize_text(text)
    if normalized in {'#e', 'e', 'n/a', 'na', 'null', 'none', 'no aplica'}:
        return ''
    return text
